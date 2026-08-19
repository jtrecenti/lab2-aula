"""Gera o dicionário da base `dados/viagens.csv.gz` em Excel.

    uv run python dicionario.py

As descrições são escritas à mão (só quem conhece o domínio sabe o que a coluna
significa); as estatísticas são calculadas do arquivo, para não envelhecerem
quando a base for regerada.
"""

from pathlib import Path

import pandas as pd

AQUI = Path(__file__).resolve().parent
BASE = AQUI / "dados" / "viagens.csv.gz"
SAIDA = AQUI / "dados" / "dicionario-viagens.xlsx"

# (coluna, papel, unidade, origem, descrição)
COLUNAS = [
    # --- identificação -----------------------------------------------------
    ("trip_id", "identificador", "", "GTFS trips.txt",
     "Viagem de referência da linha naquele sentido. No feed da SPTrans há uma viagem "
     "por linha e sentido, que serve de molde para todos os horários do dia."),
    ("route_id", "identificador / grupo", "", "GTFS routes.txt",
     "Código da linha de ônibus, no formato '1012-10'. É a coluna usada para agrupar "
     "a divisão treino e teste: nenhuma linha aparece dos dois lados."),
    ("service_id", "identificador", "", "GTFS calendar.txt",
     "Código do calendário de operação. Cada posição indica se o serviço opera em dia "
     "útil (U), sábado (S) e domingo ou feriado (D): 'U__' é só dia útil."),
    ("shape_id", "identificador", "", "GTFS trips.txt",
     "Código do traçado geográfico da viagem, usado para calcular a extensão."),
    ("stop_origem", "identificador", "", "GTFS stop_times.txt",
     "Código da primeira parada do itinerário."),
    ("stop_destino", "identificador", "", "GTFS stop_times.txt",
     "Código da última parada do itinerário."),
    ("route_short_name", "auxiliar", "", "GTFS routes.txt",
     "Código da linha como aparece no letreiro. Igual ao route_id para ônibus; para "
     "trem e metrô assume a forma 'CPTM L07' ou 'METRÔ L1'."),
    ("route_long_name", "auxiliar", "", "GTFS routes.txt",
     "Nome da linha, no formato 'origem - destino'. Serve para leitura humana e para "
     "investigar casos estranhos; não entra no modelo."),
    ("direction_id", "auxiliar", "0 ou 1", "GTFS trips.txt",
     "Sentido da viagem, na codificação do GTFS. Foi traduzido para a coluna 'sentido'."),
    ("segundos_inicio", "auxiliar", "segundos", "derivada de GTFS stop_times.txt",
     "Horário da primeira partida da viagem de referência, em segundos desde a "
     "meia-noite. Usada só para calcular a duração."),
    ("segundos_fim", "auxiliar", "segundos", "derivada de GTFS stop_times.txt",
     "Horário da chegada à última parada. Pode passar de 86400 (24h) quando a viagem "
     "vira o dia, que é como o GTFS representa isso."),

    # --- alvo ---------------------------------------------------------------
    ("headway_seg", "ALVO", "segundos", "GTFS frequencies.txt",
     "Intervalo programado entre dois ônibus da mesma linha, naquela faixa horária. "
     "É o que o modelo tenta prever. Atenção: assume apenas 13 valores distintos, de "
     "120 a 3600 segundos, porque o planejamento trabalha com intervalos redondos."),

    # --- preditoras numéricas ----------------------------------------------
    ("hora_inicio", "preditora numérica", "hora do dia (0 a 23)", "derivada de GTFS frequencies.txt",
     "Hora em que a faixa de frequência começa a valer."),
    ("n_paradas", "preditora numérica", "contagem", "derivada de GTFS stop_times.txt",
     "Número de paradas do itinerário, da origem ao destino."),
    ("extensao_km", "preditora numérica", "quilômetros", "derivada de GTFS shapes.txt",
     "Comprimento do traçado da linha. Vem do último valor acumulado de "
     "shape_dist_traveled, que o feed publica em metros."),
    ("duracao_min", "preditora numérica", "minutos", "derivada de GTFS stop_times.txt",
     "Duração programada da viagem inteira. É tempo previsto, não realizado."),
    ("velocidade_kmh", "preditora numérica", "km/h", "derivada (extensao_km / duracao_min)",
     "Velocidade comercial programada. Baixa em toda a rede porque inclui o tempo "
     "parado em cada ponto e no trânsito."),
    ("paradas_por_km", "preditora numérica", "paradas por km", "derivada (n_paradas / extensao_km)",
     "Densidade de paradas. Distingue linha troncal, com paradas espaçadas, de linha "
     "de bairro, que para a cada quarteirão."),
    ("dist_centro_origem_km", "preditora numérica", "quilômetros", "derivada de GTFS stops.txt",
     "Distância em linha reta da primeira parada até a Praça da Sé, o marco zero da "
     "cidade. Proxy de quão periférica é a ponta da linha."),
    ("dist_centro_destino_km", "preditora numérica", "quilômetros", "derivada de GTFS stops.txt",
     "Mesma medida, para a última parada."),
    ("n_linhas_origem", "preditora numérica", "contagem", "derivada de GTFS stop_times.txt",
     "Quantas linhas distintas atendem a parada de origem. Proxy de terminal ou ponto "
     "de baldeação: parada servida por muitas linhas é um nó da rede."),
    ("n_linhas_destino", "preditora numérica", "contagem", "derivada de GTFS stop_times.txt",
     "Mesma medida, para a parada de destino."),
    ("pct_paradas_corredor", "preditora numérica", "proporção de 0 a 1", "cruzamento GTFS x API Olho Vivo",
     "Fração das paradas do itinerário que ficam em corredor de ônibus. Calculada por "
     "proximidade geográfica: uma parada do GTFS a menos de 120 m de uma parada de "
     "corredor da API é contada como tal, porque os dois sistemas usam códigos de "
     "parada diferentes. Vale zero para a maioria das linhas, e zero aqui é "
     "informação, não dado faltante."),
    ("lat_origem", "auxiliar", "graus decimais", "GTFS stops.txt",
     "Latitude da primeira parada. Fora do modelo; serve para mapear resíduos."),
    ("lon_origem", "auxiliar", "graus decimais", "GTFS stops.txt", "Longitude da primeira parada."),
    ("lat_destino", "auxiliar", "graus decimais", "GTFS stops.txt", "Latitude da última parada."),
    ("lon_destino", "auxiliar", "graus decimais", "GTFS stops.txt", "Longitude da última parada."),

    # --- preditoras categóricas --------------------------------------------
    ("periodo_dia", "preditora categórica", "", "derivada de hora_inicio",
     "Faixa do dia com significado operacional: madrugada (0h a 4h), pico_manha "
     "(5h a 8h), entrepico (9h a 16h), pico_tarde (17h a 19h) e noite (20h a 23h)."),
    ("tipo_dia", "preditora categórica", "", "derivada de service_id",
     "Calendário de operação em português: todos_os_dias, util_e_sabado, "
     "fim_de_semana, domingo."),
    ("sentido", "preditora categórica", "", "derivada de direction_id",
     "'ida' ou 'volta'. A mesma linha aparece nos dois sentidos, com itinerários que "
     "podem ter tamanhos diferentes."),
    ("area_operacao", "preditora categórica", "", "derivada de route_short_name",
     "Primeiro caractere do código da linha. Os dígitos de 1 a 9 correspondem às áreas "
     "de operação da rede municipal de ônibus. As letras marcam serviços que não são "
     "ônibus municipais: N para as linhas noturnas, C para trens da CPTM e M para o "
     "Metrô, que aparecem no mesmo feed."),
    ("tipo_linha", "preditora categórica", "", "derivada de route_short_name",
     "Sufixo do código da linha, depois do hífen. O sufixo 10 é o atendimento base e "
     "responde por 88% dos registros; os demais (21, 31, 41, 51 e outros) são "
     "variações de itinerário da mesma linha. São 24 níveis na base inteira e 20 no "
     "conjunto de treino, dos quais 13 são raros, o que faz diferença no agrupamento "
     "de categorias do pré-processamento."),
    ("corredor_principal", "preditora categórica", "", "cruzamento GTFS x API Olho Vivo",
     "Nome do corredor de ônibus mais frequente entre as paradas do itinerário. "
     "Faltante em 71% dos registros, de propósito: a maioria das linhas não passa por "
     "corredor. A imputação acontece dentro do Pipeline, não aqui."),
]


def _num(valor: float, casas: int = 2) -> str:
    """Formata número no padrão brasileiro: ponto no milhar, vírgula no decimal."""
    texto = f"{valor:,.{casas}f}"
    return texto.replace(",", "M").replace(".", ",").replace("M", ".")


def montar_planilhas(dados: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """Constrói as quatro abas do dicionário."""
    documentadas = [c[0] for c in COLUNAS]
    faltando = set(dados.columns) - set(documentadas)
    if faltando:
        raise ValueError(f"colunas sem descrição no dicionário: {sorted(faltando)}")

    # aba 1: leia-me
    leiame = pd.DataFrame(
        {
            "item": [
                "Base",
                "Arquivo",
                "O que é uma linha da tabela",
                "Chave",
                "Registros",
                "Linhas de ônibus",
                "Alvo",
                "Fontes",
                "Como foi gerada",
                "Recorte temporal",
                "Cuidado 1",
                "Cuidado 2",
                "Cuidado 3",
                "Licença",
            ],
            "descrição": [
                "Oferta programada de ônibus da SPTrans",
                "dados/viagens.csv.gz",
                "Uma linha de ônibus, num sentido, numa faixa de horário do dia. "
                "A linha 8000-10, sentido ida, na faixa que começa às 7h, é um registro.",
                "route_id + sentido + hora_inicio identifica um registro (também trip_id + hora_inicio)",
                f"{len(dados):,}".replace(",", "."),
                f"{dados['route_id'].nunique():,}".replace(",", "."),
                "headway_seg, o intervalo programado entre dois ônibus, em segundos",
                "GTFS estático da SPTrans (espelho público do MobilityData) e API Olho Vivo",
                "Gerada pelo repositório github.com/jtrecenti/lab2-sptrans, com `uv run lab2 transform`",
                "Retrato do feed vigente; o GTFS é republicado periodicamente pela SPTrans",
                "É oferta PROGRAMADA, não realizada. O GTFS diz o que deveria acontecer, "
                "não o que aconteceu.",
                "Trem e metrô entram no mesmo feed (area_operacao C e M), com lógica de "
                "operação diferente da dos ônibus. Considere filtrar se isso atrapalhar.",
                "O alvo tem só 13 valores distintos, todos redondos. É uma regressão sobre "
                "uma variável bem discretizada, e isso limita o quanto o erro pode cair.",
                "Dados públicos da SPTrans. Este arquivo apenas os reorganiza.",
            ],
        }
    )

    # aba 2: dicionário de colunas
    resumo = []
    for nome, papel, unidade, origem, descricao in COLUNAS:
        serie = dados[nome]
        # identificador é código, não quantidade: mostrar faixa numérica seria enganoso
        e_numero = pd.api.types.is_numeric_dtype(serie) and "identificador" not in papel
        if e_numero:
            inteira = float(serie.dropna().mod(1).max()) == 0
            casas = 0 if inteira else 2
            faixa = f"{_num(serie.min(), casas)} a {_num(serie.max(), casas)}"
            tipico = f"mediana {_num(serie.median(), casas)}"
        else:
            faixa = f"{serie.nunique()} valores distintos"
            tipico = f"mais comum: {serie.mode().iloc[0]}" if serie.notna().any() else ""
        resumo.append(
            {
                "coluna": nome,
                "papel": papel,
                "tipo": str(serie.dtype),
                "unidade": unidade,
                "faltantes (%)": round(serie.isna().mean() * 100, 1),
                "faixa / distintos": faixa,
                "típico": tipico,
                "origem": origem,
                "descrição": descricao,
            }
        )
    dicionario = pd.DataFrame(resumo)

    # aba 3: categorias
    categorias = []
    for nome, papel, *_ in COLUNAS:
        if "categórica" not in papel and nome not in ("service_id", "tipo_linha"):
            continue
        contagem = dados[nome].value_counts(dropna=False)
        for valor, n in contagem.items():
            categorias.append(
                {
                    "coluna": nome,
                    "valor": "(faltante)" if pd.isna(valor) else valor,
                    "registros": int(n),
                    "% do total": round(n / len(dados) * 100, 2),
                }
            )
    categorias = pd.DataFrame(categorias)

    # aba 4: exemplo
    exemplo = dados.head(15).copy()

    return {
        "Leia-me": leiame,
        "Colunas": dicionario,
        "Categorias": categorias,
        "Exemplo": exemplo,
    }


def ajustar_largura(escritor: pd.ExcelWriter, planilhas: dict[str, pd.DataFrame]) -> None:
    """Deixa as colunas com largura utilizável, em vez do padrão apertado."""
    from openpyxl.utils import get_column_letter

    for aba, tabela in planilhas.items():
        planilha = escritor.sheets[aba]
        for i, coluna in enumerate(tabela.columns, start=1):
            maior = max([len(str(coluna))] + [len(str(v)) for v in tabela[coluna].head(200)])
            planilha.column_dimensions[get_column_letter(i)].width = min(maior + 2, 70)
            if maior > 70:
                from openpyxl.styles import Alignment

                for celula in planilha[get_column_letter(i)]:
                    celula.alignment = Alignment(wrap_text=True, vertical="top")
        planilha.freeze_panes = "A2"


def main() -> None:
    dados = pd.read_csv(BASE, low_memory=False)
    planilhas = montar_planilhas(dados)
    with pd.ExcelWriter(SAIDA, engine="openpyxl") as escritor:
        for aba, tabela in planilhas.items():
            tabela.to_excel(escritor, sheet_name=aba, index=False)
        ajustar_largura(escritor, planilhas)
    print(f"dicionário salvo em {SAIDA}")
    for aba, tabela in planilhas.items():
        print(f"  {aba}: {len(tabela)} linhas")


if __name__ == "__main__":
    main()
